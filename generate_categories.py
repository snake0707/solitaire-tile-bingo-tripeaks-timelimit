#!/usr/bin/env python3
"""
Generate categories.js from config xlsx files.

Reads:
  config/sort_game_category_card_config.xlsx  -- category definitions
  config/sort_game_basic_card_config.xlsx     -- card (item) definitions
  config/string_config.xlsx                   -- display name translations

Writes:
  categories.js

Usage:
  python3 generate_categories.py
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

# All three xlsx files have 5 header rows (headers, types, descriptions, scope, references)
METADATA_ROWS = 5


def load_sheet_data(filepath):
    """Load an xlsx file and return data rows (skipping metadata) as list of dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) <= METADATA_ROWS:
        return []

    headers = [str(h).strip() if h else '' for h in rows[0]]
    data = []
    for row in rows[METADATA_ROWS:]:
        entry = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                entry[h] = row[i]
        data.append(entry)
    return data


def parse_bracket_array(s):
    """Parse bracket notation like [A,B,C] into a list of strings."""
    if not s:
        return []
    s = str(s).strip()
    if s.startswith('[') and s.endswith(']'):
        inner = s[1:-1].strip()
        if not inner:
            return []
        return [item.strip() for item in inner.split(',')]
    return [s]


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    config_dir = os.path.join(base_dir, 'config')

    category_file = os.path.join(config_dir, 'sort_game_category_card_config.xlsx')
    basic_card_file = os.path.join(config_dir, 'sort_game_basic_card_config.xlsx')
    string_file = os.path.join(config_dir, 'string_config.xlsx')
    output_file = os.path.join(base_dir, 'categories.js')

    for fp in [category_file, basic_card_file, string_file]:
        if not os.path.exists(fp):
            print("Error: %s not found" % fp)
            sys.exit(1)

    # 1. Load string_config: TID -> English name
    print("Reading string_config.xlsx ...")
    string_rows = load_sheet_data(string_file)
    string_map = {}  # id -> en
    for row in string_rows:
        tid = row.get('id')
        # data_only=True resolves formulas; fall back to comments if en is None
        en = row.get('en') or row.get('comments') or ''
        if tid:
            # Replace underscores with spaces (Excel PROPER() preserves underscores)
            string_map[str(tid).strip()] = str(en).strip().replace('_', ' ')

    # 2. Load basic_card_config: card_id -> { name_tid, comments, image_res }
    print("Reading sort_game_basic_card_config.xlsx ...")
    basic_rows = load_sheet_data(basic_card_file)
    basic_map = {}  # id -> { name_tid, comments, image }
    for row in basic_rows:
        card_id = row.get('id')
        if not card_id:
            continue
        card_id = str(card_id).strip()
        name_tid = str(row.get('basic_card_name') or '').strip()
        comments = str(row.get('comments') or '').strip()
        image_res = row.get('basic_card_res')
        if image_res:
            image_res = str(image_res).strip()
        else:
            image_res = None
        basic_map[card_id] = {
            'name_tid': name_tid,
            'comments': comments,
            'image': image_res,
        }

    # 3. Load category_card_config and build CATEGORIES
    print("Reading sort_game_category_card_config.xlsx ...")
    cat_rows = load_sheet_data(category_file)

    categories = {}
    for row in cat_rows:
        cat_id = row.get('id')
        if not cat_id:
            continue
        cat_id = str(cat_id).strip()

        cat_name_tid = str(row.get('category_name') or '').strip()
        is_text_val = row.get('is_text')
        is_text = bool(int(is_text_val)) if is_text_val is not None else False

        # Resolve category display name
        cat_display_name = string_map.get(cat_name_tid, cat_name_tid)

        # Category key: text categories get "_word" suffix to avoid collisions
        cat_key = cat_display_name + '_word' if is_text else cat_display_name

        # Parse basic_card_content array
        card_ids = parse_bracket_array(row.get('basic_card_content'))

        items = []
        for cid in card_ids:
            cid = cid.strip()
            card_info = basic_map.get(cid)
            if not card_info:
                print("  Warning: card '%s' not found in basic_card_config, skipping" % cid)
                continue

            # Resolve item display name:
            # - Word cards: use TID from string_config
            # - Picture cards (no TID): use comments from basic_card_config
            if card_info['name_tid']:
                item_name = string_map.get(card_info['name_tid'], card_info['name_tid'])
            else:
                item_name = card_info['comments'].replace('_', ' ')

            # Format image path
            image = None
            if card_info['image']:
                image = "res/Item/%s.png" % card_info['image']

            # Build item id: "{Name}_{Category}" (spaces -> underscores)
            item_id = "%s_%s" % (item_name.replace(' ', '_'), cat_display_name.replace(' ', '_'))

            items.append({
                'image': image,
                'name': item_name,
                'id': item_id,
                'configId': cid,
            })

        if items:
            categories[cat_key] = {
                'isText': is_text,
                'items': items,
                'name': cat_display_name,
                'configId': cat_id,
            }
            print("  %s: %d items (isText=%s)" % (cat_key, len(items), is_text))

    # 4. Write categories.js
    print("\nWriting %s ..." % output_file)
    js_content = "// Auto-generated from xlsx configs -- do not edit manually\n"
    js_content += "// Regenerate with: python3 generate_categories.py\n"
    js_content += "const CATEGORIES = " + json.dumps(categories, indent=2, ensure_ascii=False) + ";\n\n"
    js_content += "const ALL_CATEGORY_KEYS = Object.keys(CATEGORIES);\n"

    with open(output_file, 'w') as f:
        f.write(js_content)

    print("Done! %d categories written." % len(categories))


if __name__ == '__main__':
    main()
