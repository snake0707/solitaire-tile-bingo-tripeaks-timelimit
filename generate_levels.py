#!/usr/bin/env python3
"""
Generate level JSON files from level_config.xlsx.

Reads:
  config/level_config.xlsx                    -- level definitions
  config/sort_game_category_card_config.xlsx  -- category definitions
  config/sort_game_basic_card_config.xlsx     -- card (item) definitions
  config/string_config.xlsx                   -- display name translations

Writes:
  level/level_<id>.json  (one file per level_id)

Usage:
  python3 generate_levels.py
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

# ── Level settings (translated from level_settings.js) ──────────────────────

LEVEL_SETTINGS = {
    1: {
        'bingosNeeded': 1, 'maxSlots': 3, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
        ]
    },
    2: {
        'bingosNeeded': 1, 'maxSlots': 3, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[0,1,1,1,0],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[0,1,1,1,0]],
        ]
    },
    3: {
        'bingosNeeded': 2, 'maxSlots': 3, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
        ]
    },
    4: {
        'bingosNeeded': 2, 'maxSlots': 4, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[0,0,1,0,0],[0,1,1,1,0],[1,1,1,1,1],[0,1,1,1,0],[0,0,1,0,0]],
        ]
    },
    5: {
        'bingosNeeded': 3, 'maxSlots': 4, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[0,1,1,1,0],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[0,1,1,1,0]],
        ]
    },
    6: {
        'bingosNeeded': 3, 'maxSlots': 4, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
        ]
    },
    7: {
        'bingosNeeded': 4, 'maxSlots': 5, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[0,1,1,1,0],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[0,1,1,1,0]],
        ]
    },
    8: {
        'bingosNeeded': 4, 'maxSlots': 5, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
        ]
    },
    9: {
        'bingosNeeded': 5, 'maxSlots': 5, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[0,1,1,1,0],[1,1,0,1,1],[1,0,1,0,1],[1,1,0,1,1],[0,1,1,1,0]],
            [[0,0,1,0,0],[0,1,1,1,0],[1,1,1,1,1],[0,1,1,1,0],[0,0,1,0,0]],
        ]
    },
    10: {
        'bingosNeeded': 5, 'maxSlots': 5, 'timePerCard': 10, 'penaltyTime': 5,
        'layout': [
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1]],
            [[0,1,1,1,0],[1,1,1,1,1],[1,1,1,1,1],[1,1,1,1,1],[0,1,1,1,0]],
        ]
    },
}

MAX_SETTINGS_LEVEL = max(LEVEL_SETTINGS.keys())


def count_layout_positions(layout):
    total = 0
    for layer in layout:
        for row in layer:
            for cell in row:
                total += cell
    return total


def get_level_settings(level):
    settings = LEVEL_SETTINGS.get(level, LEVEL_SETTINGS[MAX_SETTINGS_LEVEL])
    total_positions = count_layout_positions(settings['layout'])
    time_limit = total_positions * settings['timePerCard']
    return {
        'bingosNeeded': settings['bingosNeeded'],
        'maxSlots': settings['maxSlots'],
        'timeLimit': time_limit,
        'penaltyTime': settings['penaltyTime'],
    }


# ── xlsx helpers ─────────────────────────────────────────────────────────────

# Config xlsx files (string, basic_card, category_card) have 5 header rows:
# row 0 = headers, rows 1-4 = types/descriptions/scope/references
CONFIG_METADATA_ROWS = 5

# Level xlsx has 5 header rows with the same structure
LEVEL_METADATA_ROWS = 5


def load_sheet_data(filepath, skip_rows):
    """Load an xlsx file and return data rows (skipping header + metadata) as list of dicts."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) <= skip_rows:
        return []

    headers = [str(h).strip() if h else '' for h in rows[0]]
    data = []
    for row in rows[skip_rows:]:
        entry = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                entry[h] = row[i]
        data.append(entry)
    return data


# ── Bracket array parsers ────────────────────────────────────────────────────

def parse_nested_array(s):
    """Parse nested bracket notation like [[A,B],[C]] into list of lists of strings."""
    s = str(s).strip()
    if not s or s[0] != '[':
        raise ValueError('Expected "[", got: %s' % repr(s[:20]))

    if len(s) < 2 or s[1] != '[':
        raise ValueError('Expected nested array "[[", got: %s' % repr(s[:20]))

    # Remove outer brackets
    if s[-1] != ']':
        raise ValueError('Missing closing "]"')
    inner = s[1:-1]  # strip outer []

    result = []
    depth = 0
    start = -1
    for i, ch in enumerate(inner):
        if ch == '[':
            if depth == 0:
                start = i + 1
            depth += 1
        elif ch == ']':
            depth -= 1
            if depth == 0:
                segment = inner[start:i].strip()
                if segment:
                    result.append([x.strip() for x in segment.split(',')])
                else:
                    result.append([])
    return result


def parse_flat_array(s):
    """Parse flat bracket notation like [A,B,C] into list of strings."""
    s = str(s).strip()
    if not s or s[0] != '[':
        raise ValueError('Expected "[", got: %s' % repr(s[:20]))
    if s[-1] != ']':
        raise ValueError('Missing closing "]"')
    inner = s[1:-1].strip()
    if not inner:
        return []
    return [x.strip() for x in inner.split(',')]


# ── Build categories and mappings from config xlsx files ─────────────────────

def build_categories(config_dir):
    """Build categories dict and ID mappings from the three config xlsx files.

    Returns:
        categories: dict keyed by display-name-based key (e.g. 'Pets_word')
        basic_id_map: config basic card ID (e.g. 'DOG_WORD_PETS_2') -> item info
        cat_config_id_to_key: config category ID (e.g. 'WORD_PETS_2') -> cat_key
    """
    string_file = os.path.join(config_dir, 'string_config.xlsx')
    basic_card_file = os.path.join(config_dir, 'sort_game_basic_card_config.xlsx')
    category_file = os.path.join(config_dir, 'sort_game_category_card_config.xlsx')

    for fp in [string_file, basic_card_file, category_file]:
        if not os.path.exists(fp):
            print("Error: %s not found" % fp)
            sys.exit(1)

    # 1. string_config: TID -> English name
    string_rows = load_sheet_data(string_file, CONFIG_METADATA_ROWS)
    string_map = {}
    for row in string_rows:
        tid = row.get('id')
        en = row.get('en') or row.get('comments') or ''
        if tid:
            string_map[str(tid).strip()] = str(en).strip().replace('_', ' ')

    # 2. basic_card_config: card_id -> { name_tid, comments, image }
    basic_rows = load_sheet_data(basic_card_file, CONFIG_METADATA_ROWS)
    basic_map = {}
    for row in basic_rows:
        card_id = row.get('id')
        if not card_id:
            continue
        card_id = str(card_id).strip()
        name_tid = str(row.get('basic_card_name') or '').strip()
        comments = str(row.get('comments') or '').strip()
        image_res = row.get('basic_card_res')
        if image_res:
            image_res = str(image_res).strip()
        else:
            image_res = None
        basic_map[card_id] = {
            'name_tid': name_tid,
            'comments': comments,
            'image': image_res,
        }

    # 3. category_card_config -> categories dict + config ID mappings
    cat_rows = load_sheet_data(category_file, CONFIG_METADATA_ROWS)
    categories = {}
    cat_config_id_to_key = {}  # e.g. 'WORD_PETS_2' -> 'Pets_word'
    basic_id_map = {}          # e.g. 'DOG_WORD_PETS_2' -> item info

    for row in cat_rows:
        cat_id = row.get('id')
        if not cat_id:
            continue
        cat_id = str(cat_id).strip()

        cat_name_tid = str(row.get('category_name') or '').strip()
        is_text_val = row.get('is_text')
        is_text = bool(int(is_text_val)) if is_text_val is not None else False

        cat_display_name = string_map.get(cat_name_tid, cat_name_tid)
        cat_key = cat_display_name + '_word' if is_text else cat_display_name

        # Map config category ID to display key
        cat_config_id_to_key[cat_id] = cat_key

        # Parse basic_card_content array
        raw_cards = row.get('basic_card_content')
        if not raw_cards:
            card_ids = []
        else:
            raw_str = str(raw_cards).strip()
            if raw_str.startswith('[') and raw_str.endswith(']'):
                inner = raw_str[1:-1].strip()
                card_ids = [x.strip() for x in inner.split(',')] if inner else []
            else:
                card_ids = [raw_str]

        items = []
        for cid in card_ids:
            cid = cid.strip()
            card_info = basic_map.get(cid)
            if not card_info:
                continue

            if card_info['name_tid']:
                item_name = string_map.get(card_info['name_tid'], card_info['name_tid'])
            else:
                item_name = card_info['comments'].replace('_', ' ')

            image = None
            if card_info['image']:
                image = "res/Item/%s.png" % card_info['image']

            item_id = "%s_%s" % (item_name.replace(' ', '_'), cat_display_name.replace(' ', '_'))

            items.append({
                'image': image,
                'name': item_name,
                'id': item_id,
            })

            # Map config basic card ID to item info for level processing
            basic_id_map[cid] = {
                'categoryKey': cat_key,
                'name': item_name,
                'image': image,
                'isText': is_text,
            }

        if items:
            categories[cat_key] = {
                'isText': is_text,
                'items': items,
                'name': cat_display_name,
            }

    return categories, basic_id_map, cat_config_id_to_key


# ── Build one layout from a spreadsheet row ──────────────────────────────────

def build_layout(row, row_index, categories, basic_id_map, cat_config_id_to_key):
    """Convert one xlsx data row into a layout object."""
    # Parse card_content (nested array)
    raw_card = str(row.get('card_content') or '')
    try:
        card_content = parse_nested_array(raw_card)
    except ValueError as e:
        raise ValueError('Invalid card_content: %s' % e)

    # Parse category_content (flat array)
    raw_cat = str(row.get('category_content') or '')
    try:
        category_content = parse_flat_array(raw_cat)
    except ValueError as e:
        raise ValueError('Invalid category_content: %s' % e)

    if len(card_content) != 25:
        raise ValueError(
            'card_content must have 25 sub-arrays, got %d' % len(card_content))
    if len(category_content) == 0:
        raise ValueError('category_content must be a non-empty array')

    # Determine active categories from category_content
    # category_content entries are config category IDs (e.g. 'WORD_BOOK_9')
    active_cat_keys = set()
    # Resolved category entries for handPile (preserving order)
    resolved_cat_entries = []

    for cat_entry in category_content:
        # Try as config category ID first
        if cat_entry in cat_config_id_to_key:
            cat_key = cat_config_id_to_key[cat_entry]
        # Fall back to display-name-based key
        elif cat_entry in categories:
            cat_key = cat_entry
        else:
            # Try as display name
            found_key = None
            for k, v in categories.items():
                if v['name'] == cat_entry:
                    found_key = k
                    break
            if found_key:
                cat_key = found_key
            else:
                raise ValueError('Unknown category in category_content: "%s"' % cat_entry)
        active_cat_keys.add(cat_key)
        resolved_cat_entries.append(cat_key)

    # Build cards array
    cards = []
    category_targets = {}

    for i in range(25):
        sub_arr = card_content[i]
        r = i // 5
        c = i % 5

        for j, item_id in enumerate(sub_arr):
            # item_id is a config basic card ID (e.g. 'ASIA_WORD_LANDMASS_10')
            info = basic_id_map.get(item_id)
            if not info:
                raise ValueError(
                    'Unknown item ID "%s" at card_content[%d][%d]' % (item_id, i, j))

            # Layer: index 0 = topmost = highest layer number
            layer = (len(sub_arr) - 1) - j
            is_active = info['categoryKey'] in active_cat_keys
            card_type = 'regular' if is_active else 'filler'

            cards.append({
                'layer': layer,
                'row': r,
                'col': c,
                'card': {
                    'type': card_type,
                    'category': info['categoryKey'],
                    'name': info['name'],
                    'image': info['image'],
                    'isText': info['isText'],
                }
            })

            if is_active:
                category_targets[info['categoryKey']] = \
                    category_targets.get(info['categoryKey'], 0) + 1

    # Build handPile from resolved category entries
    hand_pile = []
    for cat_key in resolved_cat_entries:
        cat_obj = categories[cat_key]
        hand_pile.append({
            'type': 'gold',
            'category': cat_key,
            'name': cat_obj['name'],
            'isText': cat_obj['isText'],
        })
    hand_pile.reverse()

    return {
        'cards': cards,
        'categoryTargets': category_targets,
        'handPile': hand_pile,
        'handDisplay': [],
    }


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    config_dir = os.path.join(base_dir, 'config')
    level_file = os.path.join(config_dir, 'level_config.xlsx')
    output_dir = os.path.join(base_dir, 'level')

    if not os.path.exists(level_file):
        print("Error: %s not found" % level_file)
        sys.exit(1)

    # 1. Build categories and ID mappings
    print("Building categories and ID mappings from config files...")
    categories, basic_id_map, cat_config_id_to_key = build_categories(config_dir)
    print("  %d categories, %d basic card IDs, %d category config IDs" % (
        len(categories), len(basic_id_map), len(cat_config_id_to_key)))

    # 2. Read level config xlsx
    print("Reading %s ..." % level_file)
    rows = load_sheet_data(level_file, LEVEL_METADATA_ROWS)
    print("  %d data rows" % len(rows))

    if not rows:
        print("No data rows found.")
        return

    # 3. Group rows by level_id (preserving order)
    level_map = {}  # level_id -> [(sheet_row, row_data), ...]
    level_order = []
    for r_idx, row in enumerate(rows):
        level_id = row.get('level_id')
        sheet_row = LEVEL_METADATA_ROWS + r_idx + 1  # 1-based sheet row
        if level_id is None:
            print("  Warning: row %d missing level_id, skipping" % sheet_row)
            continue
        level_id = int(level_id)
        if level_id not in level_map:
            level_map[level_id] = []
            level_order.append(level_id)
        level_map[level_id].append((sheet_row, row))

    print("Levels to generate: %s" % ', '.join(str(x) for x in level_order))
    print()

    # 4. Convert each level
    os.makedirs(output_dir, exist_ok=True)
    error_count = 0

    for level_id in level_order:
        row_entries = level_map[level_id]
        print("-- level_%d.json (%d layout(s)) --" % (level_id, len(row_entries)))
        layouts = []
        cfg = get_level_settings(level_id)

        for sheet_row, row in row_entries:
            try:
                layout = build_layout(row, sheet_row, categories, basic_id_map, cat_config_id_to_key)

                # Per-row config: xlsx values override level_settings defaults
                row_config = {
                    'bingosNeeded': cfg['bingosNeeded'],
                    'maxSlots': cfg['maxSlots'],
                    'timeLimit': cfg['timeLimit'],
                    'penaltyTime': cfg['penaltyTime'],
                }
                bingo_num = row.get('bingo_num')
                if bingo_num is not None and bingo_num != '':
                    row_config['bingosNeeded'] = int(bingo_num)
                slot_cnt = row.get('slot_cnt')
                if slot_cnt is not None and slot_cnt != '':
                    row_config['maxSlots'] = int(slot_cnt)
                level_time = row.get('level_time')
                if level_time is not None and level_time != '':
                    row_config['timeLimit'] = int(level_time)
                deduct_time = row.get('deduct_time')
                if deduct_time is not None and deduct_time != '':
                    row_config['penaltyTime'] = int(deduct_time)

                # Build final layout with config first (matching existing JSON key order)
                final_layout = {
                    'config': row_config,
                    'cards': layout['cards'],
                    'categoryTargets': layout['categoryTargets'],
                    'handPile': layout['handPile'],
                    'handDisplay': layout['handDisplay'],
                }
                layouts.append(final_layout)

                card_count = len(layout['cards'])
                cats = ', '.join(layout['categoryTargets'].keys())
                print("  Layout row %d: %d cards, categories: %s" % (sheet_row, card_count, cats))
                print("    config: bingo=%d slots=%d time=%d penalty=%d" % (
                    row_config['bingosNeeded'], row_config['maxSlots'],
                    row_config['timeLimit'], row_config['penaltyTime']))
            except (ValueError, KeyError) as e:
                print("  Row %d ERROR: %s" % (sheet_row, e))
                error_count += 1

        if layouts:
            out_path = os.path.join(output_dir, 'level_%d.json' % level_id)
            with open(out_path, 'w', encoding='utf-8') as f:
                json.dump(layouts, f, indent=2, ensure_ascii=False)
                f.write('\n')
            print("  -> Wrote %s" % out_path)

    print()
    if error_count > 0:
        print("Done with %d error(s)." % error_count)
    else:
        print("Done. All levels converted successfully.")


if __name__ == '__main__':
    main()
